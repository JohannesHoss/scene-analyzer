from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional
import io
from datetime import datetime


class ExcelGenerator:
    """Generates formatted Excel files from analysis results"""
    
    def __init__(self, language: str, mode: str):
        self.language = language
        self.mode = mode
        self.wb = Workbook()
        self.aronson_data = None
        
    def generate(self, analysis_data: List[Dict], filename: str, aronson_data: Optional[List[Dict]] = None) -> bytes:
        """
        Generate Excel file with formatted analysis results
        
        Args:
            analysis_data: List of analyzed scene dictionaries
            filename: Original filename for title
        
        Returns:
            Excel file as bytes
        """
        ws = self.wb.active
        ws.title = "Scene Analysis"
        
        # Add title row
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = f"Scene Analysis: {filename}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        # Headers
        headers = self._get_headers()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Data rows
        for row_idx, scene in enumerate(analysis_data, 3):
            data = self._extract_data(scene)
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws, headers)
        
        # Freeze header rows
        ws.freeze_panes = "A3"
        
        # Add Aronson sheet if story mode and data available
        if "story" in self.mode and aronson_data:
            self._add_aronson_sheet(aronson_data)
        
        # Add metadata sheet if story mode
        if "story" in self.mode:
            self._add_metadata_sheet(analysis_data, filename)
        
        # Save to bytes
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _get_headers(self) -> List[str]:
        """Get column headers based on language and mode"""
        if self.language == "DE":
            headers = [
                "Szene", "INT/EXT", "Schauplatz", "Tageszeit",
                "Story Event", "Subtext", "Wendepunkt-Typ", "Wendepunkt-Moment",
                "Anwesend", "Erwähnt", "Anzahl", "Stimmung"
            ]
        else:  # EN
            headers = [
                "Scene", "INT/EXT", "Location", "Time",
                "Story Event", "Subtext", "Turn Type", "Turn Moment",
                "On Stage", "Off Stage", "Count", "Mood"
            ]
        
        # Add mode-specific columns
        if "tatort" in self.mode:
            if self.language == "DE":
                headers.extend(["Beweise", "Info-Fluss", "Wissensvorsprung", "Redundanz", "Verdächtige/Alibis"])
            else:
                headers.extend(["Evidence", "Info Flow", "Knowledge Gap", "Redundancy", "Suspects/Alibis"])
        
        if "story" in self.mode:
            if self.language == "DE":
                headers.extend(["Hero's Journey", "Akt", "Plot Point", "Erwartung"])
            else:
                headers.extend(["Hero's Journey", "Act", "Plot Point", "Expected"])
        
        return headers
    
    def _extract_data(self, scene: Dict) -> List[str]:
        """Extract data from scene dict in correct order"""
        data = [
            scene.get("number", ""),
            scene.get("int_ext", ""),
            scene.get("location", ""),
            scene.get("time_of_day", ""),
            scene.get("story_event", ""),
            scene.get("subtext", ""),
            scene.get("turning_point_type", scene.get("turning_point", "")),  # Fallback for old format
            scene.get("turning_point_moment", ""),
            ", ".join(scene.get("on_stage", [])) if isinstance(scene.get("on_stage"), list) else str(scene.get("on_stage", "")),
            ", ".join(scene.get("off_stage", [])) if isinstance(scene.get("off_stage"), list) else str(scene.get("off_stage", "")),
            len(scene.get("on_stage", [])) if isinstance(scene.get("on_stage"), list) else 0,
            scene.get("protagonist_mood", ""),
        ]
        
        # Add mode-specific data
        if "tatort" in self.mode:
            data.extend([
                scene.get("evidence", ""),
                scene.get("information_flow", ""),
                scene.get("knowledge_gap", ""),
                scene.get("redundancy", ""),
                scene.get("suspect_status", "")
            ])
        
        if "story" in self.mode:
            data.extend([
                scene.get("hero_journey", ""),
                scene.get("act", ""),
                scene.get("plot_point_actual", ""),
                scene.get("plot_point_expected", "")
            ])
        
        return data
    
    def _adjust_column_widths(self, ws, headers):
        """Auto-adjust column widths based on content"""
        column_widths = {
            0: 8,   # Scene number
            1: 10,  # INT/EXT
            2: 20,  # Location
            3: 12,  # Time
            4: 50,  # Story Event
            5: 30,  # Subtext
            6: 12,  # Turning Point Type
            7: 40,  # Turning Point Moment
            8: 25,  # On Stage
            9: 20,  # Off Stage
            10: 8,  # Count
            11: 15, # Mood
        }
        
        # Add mode-specific widths
        if "tatort" in self.mode:
            column_widths.update({
                12: 30,  # Evidence
                13: 15,  # Info Flow
                14: 18,  # Knowledge Gap
                15: 15,  # Redundancy
                16: 35   # Suspects/Alibis
            })
        
        if "story" in self.mode:
            offset = 12 if "tatort" not in self.mode else 17
            column_widths.update({
                offset: 20,     # Hero's Journey
                offset+1: 15,   # Act
                offset+2: 18,   # Plot Point
                offset+3: 20    # Expected
            })
        
        for col_idx, width in column_widths.items():
            if col_idx < len(headers):
                ws.column_dimensions[get_column_letter(col_idx + 1)].width = width
    
    def _add_aronson_sheet(self, aronson_data: List[Dict]):
        """Add Aronson analysis sheet for story mode"""
        ws = self.wb.create_sheet("Aronson Analysis", 1)  # Insert as second sheet
        
        # Title
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_text = "Aronson Single Path Analyse" if self.language == "DE" else "Aronson Single Path Analysis"
        title_cell.value = title_text
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        
        # Headers
        question_header = "Frage" if self.language == "DE" else "Question"
        answer_header = "Antwort" if self.language == "DE" else "Answer"
        
        ws['A2'] = "#"
        ws['B2'] = question_header
        ws['C2'] = answer_header
        
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}2']
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Data rows
        for idx, item in enumerate(aronson_data, 1):
            row_num = idx + 2
            
            # Number
            num_cell = ws[f'A{row_num}']
            num_cell.value = idx
            num_cell.alignment = Alignment(horizontal="center", vertical="top")
            
            # Question
            question_cell = ws[f'B{row_num}']
            question_cell.value = item.get("question", "")
            question_cell.alignment = Alignment(vertical="top", wrap_text=True)
            question_cell.font = Font(bold=True)
            
            # Answer
            answer_cell = ws[f'C{row_num}']
            answer_cell.value = item.get("answer", "")
            answer_cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Borders for all cells
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row_num}']
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Alternate row colors
                if row_num % 2 == 1:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 80
        
        # Freeze header
        ws.freeze_panes = "A3"
    
    def _add_metadata_sheet(self, analysis_data: List[Dict], filename: str):
        """Add metadata sheet for story mode"""
        ws = self.wb.create_sheet("Metadata")
        
        # Title
        ws['A1'] = "Analysis Metadata"
        ws['A1'].font = Font(size=12, bold=True)
        
        # Info
        info = [
            ("Filename", filename),
            ("Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Total Scenes", len(analysis_data)),
            ("Mode", self.mode),
            ("Language", self.language)
        ]
        
        for idx, (label, value) in enumerate(info, 3):
            ws[f'A{idx}'] = label
            ws[f'A{idx}'].font = Font(bold=True)
            ws[f'B{idx}'] = value
        
        # Auto-width
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
